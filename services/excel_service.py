from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def generate_boq_excel(boq_data, grand_total):

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Headers
    headers = ["Item No", "Description", "Unit", "Quantity", "Rate", "Amount"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Insert BOQ Rows
    for row_num, item in enumerate(boq_data, start=2):
        ws.cell(row=row_num, column=1, value=item["item_no"])
        ws.cell(row=row_num, column=2, value=item["description"])
        ws.cell(row=row_num, column=3, value=item["unit"])
        ws.cell(row=row_num, column=4, value=item["quantity"])
        ws.cell(row=row_num, column=5, value=item["rate"])
        ws.cell(row=row_num, column=6, value=item["amount"])

    # Grand Total Row
    total_row = len(boq_data) + 2

    ws.cell(row=total_row, column=5, value="Grand Total").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=grand_total).font = Font(bold=True)

    # Auto column width
    for col in ws.columns:

        max_length = 0
        column = col[0].column

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Create downloads folder automatically
    download_folder = "downloads"

    os.makedirs(download_folder, exist_ok=True)

    # Unique file name using time
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = f"boq_{timestamp}.xlsx"

    file_path = os.path.join(download_folder, file_name)

    # Save Excel
    wb.save(file_path)

    return file_path


